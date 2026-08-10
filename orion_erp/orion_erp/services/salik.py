import datetime
import io
import re

import frappe
from frappe import _
from frappe.utils import flt, get_first_day, getdate

from orion_erp.orion_erp.doctype.monthly_billing.monthly_billing import rental_on_date


def norm_plate(value) -> str:
    return re.sub(r"[\s\-]", "", str(value or "")).upper()


def parse_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    text = str(value).strip()
    try:
        return getdate(text)
    except Exception:
        pass
    try:
        from dateutil import parser as _dtparser
        return _dtparser.parse(text).date()
    except Exception:
        return None


def clean_amount(value) -> float:
    if value in (None, ""):
        return 0.0
    cleaned = re.sub(r"[^0-9.\-]", "", str(value))
    if cleaned in ("", "-", ".", "-."):
        return 0.0
    return flt(cleaned)


def cell(row, idx):
    if idx is None or idx < 0 or idx >= len(row):
        return None
    return row[idx]


def build_plate_index():
    index = {}
    for v in frappe.get_all("Vehicle", fields=["name", "license_plate", "custom_plate_code"]):
        plate = v.license_plate or v.name
        index[norm_plate(plate)] = v.name
        if v.custom_plate_code:
            index[norm_plate(f"{v.custom_plate_code}{plate}")] = v.name
    return index


@frappe.whitelist()
def import_salik(file_url, billing_month):
    frappe.has_permission("Salik or Darbs", "create", throw=True)
    month = get_first_day(getdate(billing_month))
    rows = _read_workbook(file_url)
    if not rows:
        frappe.throw(_("No toll crossings found in the attached statement."))

    plate_index = build_plate_index()
    by_vehicle, unmatched = {}, 0

    for r in rows:
        vehicle = plate_index.get(norm_plate(r.get("plate")))
        if vehicle:
            by_vehicle.setdefault(vehicle, []).append(r)
        else:
            unmatched += 1

    for vehicle, crossings in by_vehicle.items():
        _import_vehicle_crossings(vehicle, month, crossings)

    return {"vehicles": len(by_vehicle), "unmatched": unmatched}


def _import_vehicle_crossings(vehicle, month, crossings):
    doc = _get_or_create(vehicle, month)
    seen = {(getdate(c.charge_date), c.gate, flt(c.amount)) for c in (doc.crossings or [])}

    for r in crossings:
        charge_date = r.get("date") or month
        key = (getdate(charge_date), r.get("gate"), flt(r.get("amount")))
        if key in seen:
            continue
        seen.add(key)

        rental = rental_on_date(vehicle, charge_date)
        doc.append("crossings", {
            "charge_date": charge_date,
            "gate": r.get("gate"),
            "amount": flt(r.get("amount")),
            "vehicle": vehicle,
            "customer": rental.customer if rental else None,
            "rental": rental.name if rental else None,
            "company_cost": 0 if rental else 1,
        })

    doc.save(ignore_permissions=True)


def _get_or_create(vehicle, month):
    name = frappe.db.exists("Salik or Darbs", {"vehicle": vehicle, "billing_month": month, "docstatus": ["<", 2]})
    if name:
        return frappe.get_doc("Salik or Darbs", name)
    return frappe.get_doc({"doctype": "Salik or Darbs", "vehicle": vehicle, "billing_month": month})


@frappe.whitelist()
def download_template():
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Salik"
    ws.append(["Upload the Salik / Darb statement exactly as downloaded from the portal."])
    ws.append(["Transactions for Tag # 13679021 Plate # AE Abu Dhabi Private AD 22 40269"])
    for date, gate, direction, amount in [
        ("21-May-2026 02:23:02 PM", "Jebel Ali Toll Gate", "To Abu Dhabi", 4),
        ("21-May-2026 02:14:15 PM", "Al Barsha", "To Abu Dhabi", 4),
    ]:
        row = [None] * 18
        row[1], row[5], row[7], row[10], row[13], row[17] = date, "40269", "13679021", gate, direction, amount
        ws.append(row)
    ws.append(["Total transactions for Tag # 13679021"])
    buf = io.BytesIO()
    wb.save(buf)
    frappe.response["filename"] = "salik_template.xlsx"
    frappe.response["filecontent"] = buf.getvalue()
    frappe.response["type"] = "download"


def _read_workbook(file_url):
    file_name = frappe.db.get_value("File", {"file_url": file_url}, "name")
    if not file_name:
        frappe.throw(_("Could not resolve the attached file."))
    content = frappe.get_doc("File", file_name).get_content()
    return parse_salik_statement(content)


def parse_salik_statement(content):
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    out = []
    current_plate = None

    for raw in ws.iter_rows(values_only=True):
        if raw is None:
            continue

        first_cell = str(raw[0]).strip() if raw[0] is not None else ""
        current_plate = _track_plate_context(first_cell, current_plate)
        if current_plate is None and not first_cell.startswith("Transactions for Tag"):
            continue

        crossing = _parse_crossing_row(raw, current_plate)
        if crossing:
            out.append(crossing)

    return out


def _track_plate_context(first_cell, current_plate):
    if first_cell.startswith("Transactions for Tag"):
        nums = re.findall(r"\d+", first_cell)
        return nums[-1] if nums else None
    if first_cell.startswith("Total transactions"):
        return None
    return current_plate


def _parse_crossing_row(raw, current_plate):
    cells = [c for c in raw if c is not None and str(c).strip() != ""]
    if len(cells) < 2:
        return None

    charge_date = parse_date(cells[0])
    if not charge_date:
        return None

    amount = clean_amount(cells[-1])
    if amount <= 0:
        return None

    direction, gate = _extract_direction_and_gate(cells)
    return {
        "date": charge_date,
        "plate": current_plate,
        "gate": gate,
        "direction": direction,
        "amount": amount,
    }


def _extract_direction_and_gate(cells):
    middle = [str(c).strip() for c in cells[1:-1]]
    direction = next((m for m in middle if m.lower().startswith("to ")), None)
    gate = next(
        (m for m in middle if not m.lower().startswith("to ") and not m.replace(".", "").isdigit()),
        None,
    )
    return direction, gate
