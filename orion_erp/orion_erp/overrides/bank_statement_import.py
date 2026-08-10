import csv
import json
import os
import re
from urllib.parse import parse_qs, urlparse

import frappe
import openpyxl
from frappe import _
from frappe.core.doctype.data_import.importer import Importer, ImportFile
from frappe.utils.background_jobs import enqueue
from frappe.utils.xlsxutils import ILLEGAL_CHARACTERS_RE, handle_html
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from erpnext.accounts.doctype.bank_statement_import.bank_statement_import import (
    BankStatementImport as BaseBankStatementImport,
    INVALID_VALUES,
    add_bank_account,
    update_mapping_db,
)


def _is_s3_proxy_url(file_url):
    return file_url.startswith("/api/method/frappe_s3_attachment.controller.generate_file")


def resolve_file_url(file_url, doctype=None, docname=None):
    """Resolve a stored file URL to the current File doc URL.

    ``frappe_s3_attachment`` rewrites the File doc's ``file_url`` from
    ``/private/files/...`` to an S3 proxy URL, but the Bank Statement Import
    keeps the original URL. Without resolution, ``ImportFile`` cannot find the
    File doc by exact ``file_url`` match.
    """
    if not file_url:
        return file_url

    if frappe.db.exists("File", {"file_url": file_url}):
        return file_url

    if _is_s3_proxy_url(file_url):
        params = parse_qs(urlparse(file_url).query)
        key = params.get("key", [None])[0]
        if key and frappe.db.exists("File", {"content_hash": key}):
            file_url = frappe.db.get_value("File", {"content_hash": key}, "file_url")
            return file_url

        fname = params.get("file_name", [None])[0]
        if fname and frappe.db.exists("File", {"file_name": fname}):
            file_url = frappe.db.get_value("File", {"file_name": fname}, "file_url")
            return file_url

    if doctype and docname:
        for filters in (
            {
                "attached_to_doctype": doctype,
                "attached_to_name": docname,
                "attached_to_field": "import_file",
            },
            {"attached_to_doctype": doctype, "attached_to_name": docname},
        ):
            file_doc = frappe.db.get_value(
                "File", filters, ["file_url", "file_name"], as_dict=True
            )
            if file_doc and file_doc.file_url:
                return file_doc.file_url

    if file_url.startswith("/private/files/") or file_url.startswith("/files/"):
        fname = file_url.rsplit("/", 1)[-1]
        file_doc = frappe.db.get_value("File", {"file_name": fname}, ["file_url"], as_dict=True)
        if file_doc and file_doc.file_url:
            return file_doc.file_url

    return file_url


def parse_data_from_template(raw_data):
    data = []

    for _i, row in enumerate(raw_data):
        if all(v in INVALID_VALUES for v in row):
            # empty row
            continue

        data.append(row)

    return data


def write_files(import_file, data):
    full_file_path = import_file.file_doc.get_full_path()
    parts = import_file.file_doc.get_extension()
    extension = parts[1]
    extension = extension.lstrip(".")

    if not os.path.isfile(full_file_path):
        # S3-backed file: write modified data to a new local file, register it
        # as a File doc, and return its current (possibly S3-updated) URL.
        from frappe.utils.file_manager import get_files_path

        fname = frappe.utils.now_datetime().strftime("%Y%m%d%H%M%S%f") + "." + extension
        is_private = import_file.file_doc.is_private
        local_path = get_files_path(fname, is_private=is_private)
        frappe.create_folder(os.path.dirname(local_path))

        if extension == "csv":
            with open(local_path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerows(data)
        elif extension in ("xlsx", "xls"):
            write_xlsx(data, "trans", file_path=local_path)

        file_url = f"/private/files/{fname}" if is_private else f"/files/{fname}"
        new_file = frappe.get_doc(
            {
                "doctype": "File",
                "file_name": fname,
                "is_private": is_private,
                "file_url": file_url,
            }
        )
        new_file.insert(ignore_permissions=True)
        return frappe.db.get_value("File", new_file.name, "file_url") or file_url

    if extension == "csv":
        with open(full_file_path, "w", newline="") as file:
            writer = csv.writer(file)
            writer.writerows(data)
    elif extension in ("xlsx", "xls"):
        write_xlsx(data, "trans", file_path=full_file_path)

    return None


def write_xlsx(data, sheet_name, wb=None, column_widths=None, file_path=None):
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook(write_only=True)

    ws = wb.create_sheet(sheet_name, 0)

    for i, column_width in enumerate(column_widths):
        if column_width:
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width

    row1 = ws.row_dimensions[1]
    row1.font = Font(name="Calibri", bold=True)

    for row in data:
        clean_row = []
        for item in row:
            if isinstance(item, str) and (sheet_name not in ["Data Import Template", "Data Export"]):
                value = handle_html(item)
            else:
                value = item

            if isinstance(item, str) and next(ILLEGAL_CHARACTERS_RE.finditer(value), None):
                # Remove illegal characters from the string
                value = re.sub(ILLEGAL_CHARACTERS_RE, "", value)

            clean_row.append(value)

        ws.append(clean_row)

    wb.save(file_path)
    return True


def start_import(data_import, bank_account, import_file_path, google_sheets_url, bank, template_options):
    """Background job: resolve the file URL, add the bank account, then import."""
    update_mapping_db(bank, template_options)

    data_import = frappe.get_doc("Bank Statement Import", data_import)
    file = import_file_path if import_file_path else google_sheets_url

    if file and not google_sheets_url:
        file = resolve_file_url(file, "Bank Statement Import", data_import.name)

    import_file = ImportFile("Bank Transaction", file=file, import_type="Insert New Records")

    data = parse_data_from_template(import_file.raw_data)
    # Importer expects 'Data Import' class, which has 'payload_count' attribute
    if not data_import.get("payload_count"):
        data_import.payload_count = len(data) - 1

    file_path = None
    if import_file_path:
        add_bank_account(data, bank_account)
        file_path = write_files(import_file, data)

    try:
        i = Importer(data_import.reference_doctype, data_import=data_import, file_path=file_path)
        i.import_data()
    except Exception:
        frappe.db.rollback()
        data_import.db_set("status", "Error")
        data_import.log_error("Bank Statement Import failed")
    finally:
        frappe.flags.in_import = False

    # Publish only after the transaction commits, otherwise the browser's
    # refetch reads the stale status (import runs synchronously in dev mode,
    # where the commit happens after the request returns).
    frappe.publish_realtime(
        "data_import_refresh", {"data_import": data_import.name}, after_commit=True
    )


class CustomBankStatementImport(BaseBankStatementImport):
    def get_importer(self):
        if self.import_file:
            self.import_file = resolve_file_url(self.import_file, self.doctype, self.name)
        return super().get_importer()

    def start_import(self):
        preview = frappe.get_doc("Bank Statement Import", self.name).get_preview_from_template(
            self.import_file, self.google_sheets_url
        )

        if "Bank Account" not in json.dumps(preview["columns"]):
            frappe.throw(_("Please add the Bank Account column"))

        from frappe.utils.background_jobs import is_job_enqueued
        from frappe.utils.scheduler import is_scheduler_inactive

        if is_scheduler_inactive() and not frappe.flags.in_test:
            frappe.throw(_("Scheduler is inactive. Cannot import data."), title=_("Scheduler Inactive"))

        job_id = f"bank_statement_import::{self.name}"
        if not is_job_enqueued(job_id):
            enqueue(
                start_import,
                queue="default",
                timeout=6000,
                event="data_import",
                job_id=job_id,
                data_import=self.name,
                bank_account=self.bank_account,
                import_file_path=self.import_file,
                google_sheets_url=self.google_sheets_url,
                bank=self.bank,
                template_options=self.template_options,
                now=frappe.conf.developer_mode or frappe.flags.in_test,
            )
            return True

        return False
