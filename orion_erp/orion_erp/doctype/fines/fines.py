# Copyright (c) 2025, osama.ahmed@deliverydevs.com and contributors
# For license information, please see license.txt

import io

import frappe
from frappe import _
from frappe.model.document import Document
from frappe.utils import flt, getdate

from orion_erp.orion_erp.doctype.monthly_billing.monthly_billing import rental_on_date
from orion_erp.orion_erp.doctype.salik_or_darbs.salik_or_darbs import (
    _build_plate_index,
    _cell,
    _clean_amount,
    _norm_plate,
    _parse_date,
)

# Header → field aliases for the portal fines export (matched case-insensitively, exact name).
# Exact matching avoids collisions ("Fine Number"/"Plate Number" both contain "number").
_FINE_COLS = {
    "ref": ("fine number", "fine no", "reference", "ref no", "ref"),
    "status": ("status",),
    "plate": ("plate number", "plate no", "plate"),
    "date": ("date",),
    "amount": ("amount",),
    "location": ("fine location", "location"),
    "description": ("fine description", "description"),
}
_FINE_TEMPLATE_HEADERS = [
    "Fine Number",
    "Status",
    "Plate Number",
    "Date",
    "Amount",
    "Fine Location",
    "Fine Description",
]


class Fines(Document):
    def validate(self):
        # Auto-fill Customer/Driver from the vehicle's rental on the fine date.
        if self.vehicle and self.date and (not self.customer or not self.driver):
            rental = rental_on_date(self.vehicle, self.date)
            if rental:
                self.customer = self.customer or rental.customer
                if not self.driver:
                    self.driver = frappe.db.get_value("Vehicle Movement", rental.name, "driver")
        # A fine with no client rental is company-borne by default.
        if not self.closing_status and not self.customer:
            self.closing_status = "Paid by Company"

    def on_update_after_submit(self):
        if self.closing_status and self.status != "Closed":
            self.db_set("status", "Closed", update_modified=True)


@frappe.whitelist()
def create_employee_deduction(name):
    """Accounts-only: for a Driver-responsible fine, create a DRAFT Employee Deduction."""
    frappe.only_for(("Accounts Manager", "System Manager"))
    doc = frappe.get_doc("Fines", name)
    if doc.closing_status != "Paid by Driver":
        frappe.throw(_("Employee deduction applies only when Responsibility is 'Paid by Driver'."))
    if doc.employee_deduction:
        frappe.throw(_("An Employee Deduction already exists: {0}").format(doc.employee_deduction))
    if not doc.driver:
        frappe.throw(_("Set the Driver first."))
    employee = frappe.db.get_value("Driver", doc.driver, "employee")
    if not employee:
        frappe.throw(_("Driver {0} has no linked Employee.").format(doc.driver))

    deduction = frappe.get_doc(
        {
            "doctype": "Employee Deduction",
            "naming_series": "EMP-DED-.YYYY.-",
            "employee": employee,
            "employee_name": frappe.db.get_value("Employee", employee, "employee_name"),
            "transaction_date": doc.date,
            "remarks": _("Fine {0} — Vehicle {1} — AED {2}. Add penalty rows and submit.").format(
                doc.name, doc.vehicle or "-", flt(doc.amount)
            ),
        }
    )
    deduction.insert(ignore_permissions=True)
    doc.db_set("employee_deduction", deduction.name, update_modified=True)
    return deduction.name


# ---------------------------------------------------------------------------
# Bulk import — one portal fines export → one Fine document per row
# ---------------------------------------------------------------------------
@frappe.whitelist()
def import_fines(file_url):
    frappe.has_permission("Fines", "create", throw=True)
    rows = _read_fine_workbook(file_url)
    if not rows:
        frappe.throw(_("No data rows found in the attached Excel."))
    plate_index = _build_plate_index()
    created, unmatched, duplicates = 0, 0, 0
    for r in rows:
        vehicle = plate_index.get(_norm_plate(r.get("plate")))
        if not vehicle:
            unmatched += 1
            continue
        ref = str(r.get("ref")).strip() if r.get("ref") not in (None, "") else None
        if ref and frappe.db.exists("Fines", {"ref_no": ref, "docstatus": ["<", 2]}):
            duplicates += 1
            continue
        date = _parse_date(r.get("date")) or getdate(frappe.utils.nowdate())
        rental = rental_on_date(vehicle, date)
        status = str(r.get("status")).strip() if r.get("status") not in (None, "") else None
        doc = frappe.get_doc(
            {
                "doctype": "Fines",
                "date": date,
                "vehicle": vehicle,
                "amount": _clean_amount(r.get("amount")),
                "ref_no": ref,
                "portal_status": status,
                "fine_description": r.get("description"),
                "fine_location": r.get("location"),
                "customer": rental.customer if rental else None,
                "driver": frappe.db.get_value("Vehicle Movement", rental.name, "driver") if rental else None,
                "closing_status": "Paid by Client" if rental else "Paid by Company",
            }
        )
        doc.insert(ignore_permissions=True)
        doc.submit()
        created += 1
    return {"created": created, "unmatched": unmatched, "duplicates": duplicates}


@frappe.whitelist()
def download_template():
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fines"
    ws.append(_FINE_TEMPLATE_HEADERS)
    buf = io.BytesIO()
    wb.save(buf)
    frappe.response["filename"] = "fines_template.xlsx"
    frappe.response["filecontent"] = buf.getvalue()
    frappe.response["type"] = "download"


def _read_fine_workbook(file_url):
    file_name = frappe.db.get_value("File", {"file_url": file_url}, "name")
    if not file_name:
        frappe.throw(_("Could not resolve the attached file."))
    content = frappe.get_doc("File", file_name).get_content()
    return _parse_fine_statement(content)


def _parse_fine_statement(content):
    """Parse the flat portal fines export into row dicts, mapping by exact header name."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    col = _fine_columns(header)
    out = []
    for raw in rows[1:]:
        if raw is None or all(c is None for c in raw):
            continue
        plate = _cell(raw, col.get("plate"))
        if plate in (None, ""):
            continue
        out.append(
            {
                "ref": _cell(raw, col.get("ref")),
                "status": _cell(raw, col.get("status")),
                "plate": plate,
                "date": _cell(raw, col.get("date")),
                "amount": _cell(raw, col.get("amount")),
                "location": _cell(raw, col.get("location")),
                "description": _cell(raw, col.get("description")),
            }
        )
    return out


def _fine_columns(header):
    pos = {}
    for i, h in enumerate(header):
        pos.setdefault(h, i)
    return {key: next((pos[a] for a in aliases if a in pos), None) for key, aliases in _FINE_COLS.items()}
