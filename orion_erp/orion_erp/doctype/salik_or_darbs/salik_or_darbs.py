# Copyright (c) 2025, osama.ahmed@deliverydevs.com and contributors
# For license information, please see license.txt

import datetime
import io
import re

import frappe
from frappe import _
from frappe.model.document import Document
from frappe.utils import flt, get_first_day, getdate

from orion_erp.orion_erp.doctype.monthly_billing.monthly_billing import rental_on_date


def _norm_plate(value) -> str:
    return re.sub(r"[\s\-]", "", str(value or "")).upper()


def _parse_date(value):
    """Tolerant date parse for portal cells (datetime objects, ISO, or '21-May-2026 02:23:02 PM')."""
    if value in (None, ""):
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    text = str(value).strip()
    try:
        return getdate(text)
    except Exception:
        pass
    try:
        from dateutil import parser as _dtparser

        return _dtparser.parse(text).date()
    except Exception:
        return None


def _clean_amount(value) -> float:
    """Strip currency formatting like ' AED 1,000.00 ' down to a float."""
    if value in (None, ""):
        return 0.0
    cleaned = re.sub(r"[^0-9.\-]", "", str(value))
    if cleaned in ("", "-", ".", "-."):
        return 0.0
    return flt(cleaned)


class SalikorDarbs(Document):
    def autoname(self):
        self.billing_month = get_first_day(getdate(self.billing_month))
        self.name = f"SAL-{self.vehicle}-{self.billing_month}"

    def validate(self):
        if self.billing_month:
            self.billing_month = get_first_day(getdate(self.billing_month))
        self.total_amount = sum(flt(r.amount) for r in (self.crossings or []))


# ---------------------------------------------------------------------------
# Bulk import — one monthly Salik/Darb statement → one Salik doc per vehicle
# ---------------------------------------------------------------------------
@frappe.whitelist()
def import_salik(file_url, billing_month):
    frappe.has_permission("Salik or Darbs", "create", throw=True)
    month = get_first_day(getdate(billing_month))
    rows = _read_workbook(file_url)
    if not rows:
        frappe.throw(_("No toll crossings found in the attached statement."))

    plate_index = _build_plate_index()
    by_vehicle, unmatched = {}, 0
    for r in rows:
        vehicle = plate_index.get(_norm_plate(r.get("plate")))
        if vehicle:
            by_vehicle.setdefault(vehicle, []).append(r)
        else:
            unmatched += 1

    for vehicle, crossings in by_vehicle.items():
        doc = _get_or_create(vehicle, month)
        # Idempotent re-import: don't re-append a crossing already on the doc.
        seen = {(getdate(c.charge_date), c.gate, flt(c.amount)) for c in (doc.crossings or [])}
        for r in crossings:
            charge_date = r.get("date") or month
            key = (getdate(charge_date), r.get("gate"), flt(r.get("amount")))
            if key in seen:
                continue
            seen.add(key)
            rental = rental_on_date(vehicle, charge_date)
            doc.append(
                "crossings",
                {
                    "charge_date": charge_date,
                    "gate": r.get("gate"),
                    "amount": flt(r.get("amount")),
                    "vehicle": vehicle,
                    "customer": rental.customer if rental else None,
                    "rental": rental.name if rental else None,
                    "company_cost": 0 if rental else 1,
                },
            )
        doc.save(ignore_permissions=True)

    return {"vehicles": len(by_vehicle), "unmatched": unmatched}


def _get_or_create(vehicle, month):
    name = frappe.db.exists("Salik or Darbs", {"vehicle": vehicle, "billing_month": month, "docstatus": ["<", 2]})
    if name:
        return frappe.get_doc("Salik or Darbs", name)
    return frappe.get_doc({"doctype": "Salik or Darbs", "vehicle": vehicle, "billing_month": month})


@frappe.whitelist()
def download_template():
    """Emit a short sample mirroring the portal's sectioned layout (real use: upload the raw export)."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Salik"
    ws.append(["Upload the Salik / Darb statement exactly as downloaded from the portal."])
    ws.append(["Transactions for Tag # 13679021 Plate # AE Abu Dhabi Private AD 22 40269"])
    for date, gate, direction, amount in [
        ("21-May-2026 02:23:02 PM", "Jebel Ali Toll Gate", "To Abu Dhabi", 4),
        ("21-May-2026 02:14:15 PM", "Al Barsha", "To Abu Dhabi", 4),
    ]:
        row = [None] * 18
        row[1], row[5], row[7], row[10], row[13], row[17] = date, "40269", "13679021", gate, direction, amount
        ws.append(row)
    ws.append(["Total transactions for Tag # 13679021"])
    buf = io.BytesIO()
    wb.save(buf)
    frappe.response["filename"] = "salik_template.xlsx"
    frappe.response["filecontent"] = buf.getvalue()
    frappe.response["type"] = "download"


# ---------------------------------------------------------------------------
# Statement parsing
# ---------------------------------------------------------------------------
def _read_workbook(file_url):
    file_name = frappe.db.get_value("File", {"file_url": file_url}, "name")
    if not file_name:
        frappe.throw(_("Could not resolve the attached file."))
    content = frappe.get_doc("File", file_name).get_content()
    return _parse_salik_statement(content)


def _parse_salik_statement(content):
    """Parse the sectioned portal statement into flat crossing dicts.

    The export groups transactions under per-vehicle headers
    ``Transactions for Tag # <tag> Plate # … <plate#>``. Within a transaction row the
    fields always appear in the order [date, plate, tag, gate, direction, amount], but the
    portal shifts the leading column offset between sections — so we map by POSITION among the
    row's non-empty cells (date = first, amount = last, direction = the "To …" cell, gate = the
    remaining text) rather than by absolute column index. Account/Summary/Payment sections
    (before any tag header) are skipped automatically.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    out = []
    current_plate = None
    for raw in ws.iter_rows(values_only=True):
        if raw is None:
            continue
        c0 = str(raw[0]).strip() if raw[0] is not None else ""
        if c0.startswith("Transactions for Tag"):
            nums = re.findall(r"\d+", c0)
            current_plate = nums[-1] if nums else None
            continue
        if c0.startswith("Total transactions"):
            current_plate = None
            continue
        if current_plate is None:
            continue  # inside Summary / Payment / Adjustment — not a crossing

        cells = [c for c in raw if c is not None and str(c).strip() != ""]
        if len(cells) < 2:
            continue
        charge_date = _parse_date(cells[0])  # the transaction date is always the first value
        if not charge_date:
            continue
        amount = _clean_amount(cells[-1])  # the toll amount is always the rightmost value
        if amount <= 0:
            continue  # skip AED 0 free crossings (decision: charged rows only)

        middle = [str(c).strip() for c in cells[1:-1]]  # plate, tag, gate, direction
        direction = next((m for m in middle if m.lower().startswith("to ")), None)
        gate = next(
            (m for m in middle if not m.lower().startswith("to ") and not m.replace(".", "").isdigit()),
            None,
        )
        out.append(
            {
                "date": charge_date,
                "plate": current_plate,
                "gate": gate,
                "direction": direction,
                "amount": amount,
            }
        )
    return out


def _cell(row, idx):
    if idx is None or idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _build_plate_index():
    index = {}
    for v in frappe.get_all("Vehicle", fields=["name", "license_plate", "custom_plate_code"]):
        plate = v.license_plate or v.name
        index[_norm_plate(plate)] = v.name
        if v.custom_plate_code:
            index[_norm_plate(f"{v.custom_plate_code}{plate}")] = v.name
    return index
